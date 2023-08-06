import mysql.connector  # pip install mysql-connector-python
import openpyxl  # pip install openpyxl


def dump():
    cursor_created = False
    connection_established = False
    wb = openpyxl.Workbook()
    sheet = wb.active
    un = input('Enter USERNAME to connect MySQL database:')
    pw = input(f'Enter PASSWORD of user {un}:')
    db = input('Enter name of DATABASE to access:')
    tbl = input(f'Enter name of TABLE in database {db}:')
    try:
        conn = mysql.connector.connect(user=un, password=pw, host='localhost', database=db,
                                       auth_plugin='mysql_native_password')
        connection_established = True

        curs = conn.cursor()
        cursor_created = True

        curs.execute(f"DESC {tbl}")
        column_no = 1
        for db_row in curs:
            sheet.cell(row=1, column=column_no).value = db_row[0]
            column_no += 1

        curs.execute(f"SELECT * FROM {tbl}")
        row_no = 2
        for db_row in curs:
            column_no = 1
            for db_column in db_row:
                sheet.cell(row=row_no, column=column_no).value = db_column
                column_no += 1
            row_no += 1
        path = f'{tbl}.xlsx'
        wb.save(path)

        conn.commit()

    except mysql.connector.errors.ProgrammingError as mcePE:
        print('ProgrammingError', mcePE)
        if mcePE.errno == 1045:
            print('Please check your username and password passed in connect method!')
        elif mcePE.errno == 1049:
            print('Please check your database name passed in connect method!')
        elif mcePE.errno == 1064:
            print('Please check your SQL query passed in execute method!')
        elif mcePE.errno == 1146:
            print('Please check name of the table in SQL query passed to execute method!')

    except mysql.connector.errors.DatabaseError as mceDE:
        print(mceDE, '\nOR Please check your host passed in connect method!')

    except Exception as e:
        print(e)

    else:
        print(f'\nDone! Data of "{tbl}" table from "{db}" database successfully exported to excel sheet\
        "{path}" in current directory!')

    finally:
        if cursor_created:
            curs.close()
            print('\nCursor Closed!')
        if connection_established:
            conn.close()
            print('Connection Closed!')
